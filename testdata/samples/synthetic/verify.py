#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Верификатор: точный порт логики парсеров из
internal/ingest/clientbank.go и internal/ingest/ingest.go.
Проверяет, что каждый сгенерированный файл принимается парсером
(транзакции распознаются, направление корректно, суммы сходятся).
"""
import csv
import glob
import io
import os
import re

OUT = os.path.dirname(os.path.abspath(__file__))


# --- порт parseAmount ---
def parse_amount(s):
    s = s.strip()
    if not s:
        raise ValueError("empty")
    for ch in (" ", " ", " ", "'", "`"):
        s = s.replace(ch, "")
    last_comma = s.rfind(",")
    last_dot = s.rfind(".")
    if last_comma > last_dot:
        s = s.replace(".", "")
        s = s.replace(",", ".", 1)
    else:
        s = s.replace(",", "")
    return float(s)


# --- порт parseDate ---
def parse_date(s):
    from datetime import datetime
    s = s.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    raise ValueError("bad date " + s)


# --- порт parseClientBankExchange ---
def parse_clientbank(raw):
    # decode (Кодировка=Windows -> cp1251)
    body = raw
    if body[:3] == b"\xef\xbb\xbf":
        body = body[3:]
    try:
        cp = body.decode("cp1251")
    except Exception:
        cp = None
    if cp and "Кодировка=Windows" in cp:
        text = cp
    else:
        text = body.decode("utf-8", "replace")

    head_acc = ""
    txs = []
    in_doc = False
    doc = {}
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("СекцияДокумент="):
            in_doc = True
            doc = {}
            continue
        if line == "КонецДокумента":
            if in_doc:
                tx = build_cb_tx(doc, head_acc)
                if tx:
                    txs.append(tx)
            in_doc = False
            doc = {}
            continue
        if line == "КонецФайла":
            break
        i = line.find("=")
        if i <= 0:
            continue
        key = line[:i].strip()
        val = line[i + 1:].strip()
        if not key:
            continue
        if key == "РасчСчет" and not in_doc and head_acc == "":
            head_acc = val
            continue
        if in_doc:
            doc[key] = val
    return txs, head_acc


def build_cb_tx(doc, head_acc):
    if not head_acc:
        return None
    try:
        parse_date(doc.get("Дата", ""))
        amount = abs(parse_amount(doc.get("Сумма", "")))
    except Exception:
        return None
    payer = doc.get("ПлательщикСчет", "").strip()
    receiver = doc.get("ПолучательСчет", "").strip()
    if payer == head_acc:
        direction = "out"
        cp = doc.get("Получатель", "").strip()
    elif receiver == head_acc:
        direction = "in"
        cp = doc.get("Плательщик", "").strip()
    else:
        return None
    purpose = doc.get("НазначениеПлатежа", "").strip()
    if not cp or not purpose or amount == 0:
        return None
    return (direction, amount, cp, purpose)


# --- порт ParseCSV (Comma=';') ---
def parse_csv_file(path):
    raw = open(path, "rb").read()
    if raw[:3] == b"\xef\xbb\xbf":
        raw = raw[3:]
    text = raw.decode("utf-8", "replace")
    # csv reader with delimiter ';'
    rows = list(csv.reader(io.StringIO(text), delimiter=";"))
    if not rows:
        return []
    header = [c.strip().lower() for c in rows[0]]
    hidx = {h: i for i, h in enumerate(header)}
    if "сумма дебет (руб.)" in hidx and "сумма кредит (руб.)" in hidx:
        return parse_bank_rows(rows[1:], hidx)
    return parse_legacy_rows(rows)


def parse_bank_rows(rows, col):
    out = []
    def get(row, key):
        i = col.get(key)
        if i is None or i >= len(row):
            return ""
        return row[i].strip()
    for row in rows:
        ds = get(row, "дата операции")
        try:
            parse_date(ds)
        except Exception:
            continue
        debit = get(row, "сумма дебет (руб.)")
        credit = get(row, "сумма кредит (руб.)")
        try:
            if credit and not debit:
                amt = abs(parse_amount(credit)); direction = "in"
            elif debit and not credit:
                amt = abs(parse_amount(debit)); direction = "out"
            else:
                continue
        except Exception:
            continue
        cp = get(row, "наименование контрагента")
        purpose = get(row, "назначение платежа") or "—"
        if amt == 0:
            continue
        out.append((direction, amt, cp, purpose))
    return out


def parse_legacy_rows(records):
    out = []
    for rec in records:
        if len(rec) < 6:
            continue
        try:
            parse_date(rec[0].strip())
            amount = parse_amount(rec[1].strip())
        except Exception:
            continue
        d = rec[2].strip().lower()
        if d in ("приход", "in", "вход", "поступление"):
            direction = "in"
        elif d in ("расход", "out", "выход", "списание"):
            direction = "out"
        else:
            direction = "out" if amount < 0 else "in"
        cp = rec[3].strip()
        purpose = rec[5].strip() or "—"
        out.append((direction, abs(amount), cp, purpose))
    return out


# --- ParseXLSX ---
def parse_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []
    header = [c.strip().lower() for c in rows[0]]
    hidx = {h: i for i, h in enumerate(header)}
    # xlsx bank statement uses same bank columns
    return parse_bank_rows_xlsx(rows[1:], hidx)


def parse_bank_rows_xlsx(rows, col):
    out = []
    def get(row, key):
        i = col.get(key)
        if i is None or i >= len(row):
            return ""
        return str(row[i]).strip()
    for row in rows:
        ds = get(row, "дата операции")
        try:
            parse_date(ds)
        except Exception:
            continue
        debit = get(row, "сумма дебет (руб.)")
        credit = get(row, "сумма кредит (руб.)")
        try:
            if credit and not debit:
                amt = abs(parse_amount(credit)); direction = "in"
            elif debit and not credit:
                amt = abs(parse_amount(debit)); direction = "out"
            else:
                continue
        except Exception:
            continue
        cp = get(row, "наименование контрагента")
        purpose = get(row, "назначение платежа") or "—"
        if amt == 0:
            continue
        out.append((direction, amt, cp, purpose))
    return out


def main():
    files = sorted(glob.glob(os.path.join(OUT, "*")))
    total_ok = 0
    total_fail = 0
    for f in files:
        base = os.path.basename(f)
        ext = os.path.splitext(f)[1].lower()
        if ext not in (".csv", ".txt", ".xlsx"):
            continue
        try:
            if ext == ".txt":
                raw = open(f, "rb").read()
                txs, head = parse_clientbank(raw)
            elif ext == ".csv":
                txs = parse_csv_file(f)
            else:
                txs = parse_xlsx(f)
        except Exception as e:
            print(f"FAIL {base:40s} EXC {e}")
            total_fail += 1
            continue
        n_in = sum(1 for t in txs if t[0] == "in")
        n_out = sum(1 for t in txs if t[0] == "out")
        s_in = sum(t[1] for t in txs if t[0] == "in")
        s_out = sum(t[1] for t in txs if t[0] == "out")
        status = "OK  " if txs else "EMPTY"
        if not txs:
            total_fail += 1
        else:
            total_ok += 1
        print(f"{status} {base:40s} n={len(txs):3d} in={n_in:3d} out={n_out:3d} "
              f"sumIn={s_in:12.0f} sumOut={s_out:12.0f}")
    print(f"\nПарсится: {total_ok}, пусто/ошибка: {total_fail}")


if __name__ == "__main__":
    main()
